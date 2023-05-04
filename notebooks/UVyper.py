import pandas as pd
from pandas import DataFrame
from vyper.user import Model
from vyper.utils.tools import StatisticalTools as st
from vyper.user.explorer import DataProfiler
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, colors
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image

import math
import numpy as np
import pickle
import os
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.io as pio
from vyper.utils.tools import StatisticalTools as st
from sklearn.preprocessing import OrdinalEncoder
import scipy as stats
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from scipy.stats import chi2, chi2_contingency
from varclushi import VarClusHi

from scipy.spatial.distance import cdist
from sklearn.preprocessing import LabelEncoder
from k_means_constrained import KMeansConstrained
from sklearn.decomposition import PCA
from sklearn.metrics import silhouette_score, davies_bouldin_score
from kneed import KneeLocator
from yellowbrick.cluster import KElbowVisualizer
from matplotlib.cm import viridis
import scipy.cluster.hierarchy as shc
from sklearn.cluster import AgglomerativeClustering, DBSCAN, Birch, KMeans
from sklearn.mixture import GaussianMixture
from sklearn.model_selection import GridSearchCV, RandomizedSearchCV
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier

thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")


class Preprocessing:
    def __init__(self, data: str):
        """
        __init__ method for the Preprocessing class
        :param data: string - path to the data
        """
        self.df = pd.read_csv(data)

    def get_df(self):
        """
        returns the data
        :return: dataframe
        """
        return self.df

    def get_shape(self):
        """
        prints the shape of the data
        :return: tuple
        """
        return self.df.shape

    def vyper(self, dependent_variable: str):
        """
        Method to insert data into vyper
        :param dependent_variable: string
        :return: vyper model
        """
        m = Model(
            data=self.df,
            dependent_variable=dependent_variable,
            na_drop_threshold=0.5,  # set here when variables should be dropped 1 = none dropped
            training_percentage=0.7,  # set here % of dataset that should be training
            model_type="linear",
        )
        return m

    @staticmethod
    def show_variable_types(m):
        """
        Method to show the variable types
        :param m: vyper model
        :return: returns the variable types
        """
        return m.variables.show_types()

    @staticmethod
    def quantile_function(a: list, bounds: float):
        """
        Method to calculate the quantile function
        :param a: list - a list of variables for which the quantile function is to be calculated
        :param bounds: float - a float representing the bounds around the median for which the quantile function is to be calculated
        :return: list - returns the lower and upper bound
        """
        lower_bound = np.quantile(a, q=(0.5 - bounds / 2), interpolation="nearest")
        upper_bound = np.quantile(a, q=(0.5 + bounds / 2), interpolation="nearest")
        final_bounds = [lower_bound, upper_bound]

        return final_bounds

    def recoding(self, bounds: float, min_bin_size: float, dependent_variable: str,
                 ordinal_variables: list = None):
        """
        Method to recode the data
        :param bounds: float - this parameter is used to calculate the quantile function for numeric variables. It determines the range of the bin where the data will be recoded.
        :param min_bin_size: float - this parameter is used to determine whether or not a binary variable should be split into two separate variables. If the proportion of the most frequent value in the binary variable is less than min_bin_size, the variable is left as is. Otherwise, two variables are created: one indicating the presence of the most frequent value, and another indicating the missing values.
        :param dependent_variable: string - this parameter specifies the dependent variable of the dataset. This variable is used to train a vyper model to determine which variables should be recoded.
        :param ordinal_variables: list - this parameter is used to specify which variables should be treated as ordinal variables. If a variable is specified as ordinal, it will be recoded using factorization. If it is not specified as ordinal, it will be treated as a numeric variable and recoded using the quantile function.
        :return: dataframe, list, list, set, list
        """
        m = self.vyper(dependent_variable)
        original_variables = m.data.columns.to_list()
        excluded_variables = m.variables.get_excluded_variables()
        category_variables = m.variables.get_categorical_variables()
        numeric_variables = m.variables.get_numeric_variables()
        binary_variables = m.variables.get_binary_variables()
        keep_variables = original_variables
        if ordinal_variables is None:
            ordinal_variables = []
        else:
            ordinal_variables = ordinal_variables
            for var in ordinal_variables:
                if var in excluded_variables:
                    excluded_variables.remove(var)
                elif var in category_variables:
                    category_variables.remove(var)
                elif var in numeric_variables:
                    numeric_variables.remove(var)
                elif var in binary_variables:
                    binary_variables.remove(var)

        nv = list(numeric_variables)
        bv = list(binary_variables)
        ov = list(ordinal_variables)

        for var in excluded_variables:
            if var in self.df.columns:
                keep_variables.remove(var)
                self.df.drop(var, axis=1, inplace=True)
        print("Excluded variables (vyper): ", excluded_variables)
        for var in numeric_variables:
            if var in self.df.columns:
                Q3 = np.quantile(self.df[[var]], 0.75)
                Q1 = np.quantile(self.df[[var]], 0.25)
                IQR = Q3 - Q1

                if IQR == 0:
                    keep_variables.remove(var)
                    nv.remove(var)
                    self.df.drop(var, axis=1, inplace=True)
                    print(
                        "The following variables is excluded due to inter quantile equal 0:"
                        + var
                    )

                else:
                    bnds = self.quantile_function(
                        [var for var in list(self.df[var]) if not math.isnan(var)],
                        bounds=bounds,
                    )
                    self.df[var + "_processed"] = pd.Series(
                        np.minimum(np.maximum(self.df[var], bnds[0]), bnds[1])
                    )
                    keep_variables.remove(var)
                    nv.remove(var)
                    keep_variables.append(var + "_processed")
                    nv.append(var + "_processed")
                    self.df = self.df[keep_variables]

        for var in ordinal_variables:
            if var in self.df.columns:
                training_data = self.df[[var]].drop_duplicates()
                training_data[var + "_processed"] = (
                        pd.factorize(training_data[var], sort=True)[0] + 1
                )
                self.df = self.df.merge(training_data, on=var)
                keep_variables.remove(var)
                ov.remove(var)
                keep_variables.append(var + "_processed")
                ov.append(var + "_processed")
                self.df = self.df[keep_variables]

        N = self.df.shape[0]

        for var in binary_variables:
            if var in self.df.columns:
                tab = self.df[var].value_counts(ascending=False)
                counter2 = 0
                counter1 = tab.iloc[1]
                counter2 = counter1 + counter2
                N_missing = sum(self.df[var].isna())

                if N_missing / N >= min_bin_size:
                    self.df[var + "_missing_ind"] = np.where(
                        self.df[var].isna() == 1, 1, 0
                    )

                if counter2 / (N - N_missing) >= min_bin_size:
                    self.df[var + "_ind"] = np.where(self.df[var] == tab.index[1], 1, 0)
                self.df = self.df.drop([var], axis=1)
                bv.remove(var)
                bv.append(var + "_ind")

        return self.df, nv, bv, category_variables, ov

    def category_encoding(self, category_variables: list):
        """
        Method to encode the categorical variables
        :param category_variables: list - a list of categorical variables in the DataFrame
        """
        if not category_variables:
            print("No categorical variables found to encode")
            return "No categorical variables found to encode"
        temp = self.df[category_variables]
        temp = pd.get_dummies(
            temp, prefix=category_variables, columns=category_variables
        )
        self.df.drop(category_variables, axis=1, inplace=True)
        self.df = pd.concat([self.df, temp], axis=1)

    def missing_zero_values_table(self):
        """
        Method to generate a table that represents the number and percentage of missing and zero values for each column in the dataset. The table includes columns for the number of zero values, number of missing values, percentage of missing values, total number of zero and missing values, percentage of total zero and missing values, and data type.
        :return: Pandas DataFrame containing the missing and zero values table.
        """

        zero_val = (self.df == 0.00).astype(int).sum(axis=0)
        mis_val = self.df.isnull().sum()
        mis_val_percent = (self.df.isnull().sum() / len(self.df)) * 100
        missing_table = pd.concat([zero_val, mis_val, mis_val_percent], axis=1)
        missing_table = missing_table.rename(
            columns={0: "Zero Values", 1: "Missing Values", 2: "% of Total Values"}
        )
        missing_table["Total Zero + Missing Values"] = (
                missing_table["Zero Values"] + missing_table["Missing Values"]
        )
        missing_table["% Total Zero + Missing Values"] = (missing_table["Total Zero + Missing Values"] / len(
            self.df)) * 100
        missing_table = (
            missing_table[missing_table.iloc[:, 1] != 0]
            .sort_values("% of Total Values", ascending=False)
            .round(1)
        )
        missing_table.reset_index(inplace=True)
        missing_table.rename(columns={'index': 'feature'}, inplace=True)
        missing_table['Data type'] = missing_table['feature'].apply(lambda feature: str(self.df[feature].dtype))
        return missing_table

    def drop_missing(self, threshold: float):
        """
        Method to drop the columns with missing values greater than the threshold
        :param threshold: float - specifies the threshold value for the proportion of missing values in a column. Columns with missing values greater than this threshold will be dropped from the dataset.
        """
        drop_list = []
        for cols in self.df.columns:
            if (self.df.loc[:, cols].isnull().sum() / len(self.df)) > int(threshold):
                drop_list.append(cols)
        self.df.drop(drop_list, axis=1, inplace=True)

    def impute_na(self, columns_list: list, method: str):  # mean, mode,  bfill, ffill
        """
        Method to impute the missing values
        :param columns_list: list -  a list of columns to impute missing values
        :param method: string -  a string specifying the imputation method, which can be one of the following:
                                    "mean": imputes missing values with the mean value of the column
                                    "mode": imputes missing values with the mode value of the column
                                    "bfill": imputes missing values with the next valid value in the column (backward fill)
                                    "ffill": imputes missing values with the previous valid value in the column (forward fill)
        """
        if method == "mean":
            for i in columns_list:
                if i in self.df.columns:
                    self.df[i].fillna(self.df[i].mean(), inplace=True)
        if method == "mode":
            for i in columns_list:
                if i in self.df.columns:
                    self.df[i].fillna(self.df[i].mode()[0], inplace=True)

        if method == "ffill" or method == "bfill":
            for i in columns_list:
                if i in self.df.columns:
                    self.df[i].fillna(method=method, inplace=True)

    def correlation(self, numerical_variables: list, ordinal_variables: list = None, threshold: float = 0.8):
        """
        Method to remove the highly correlated variables
        :param numerical_variables: list - list of numerical variables to consider for correlation analysis.
        :param ordinal_variables: list - list of ordinal variables to consider for correlation analysis.
        :param threshold: float - threshold value for correlation. Variables with correlation greater than this threshold will be removed from the dataset.
        :return: dataframe
        """
        if ordinal_variables is None:
            ordinal_variables = []
        df_copy = self.df.copy()
        cols = self.df.columns
        newcol = []
        for col in numerical_variables + ordinal_variables:
            if col in cols:
                newcol.append(col)
        df_copy = df_copy[newcol]
        list1 = list()
        col_corr = set()  # Set of all the names of deleted columns
        corr_matrix = df_copy.corr()
        for i in range(len(corr_matrix.columns)):
            for j in range(i):
                if (corr_matrix.iloc[i, j] >= threshold) and (
                        corr_matrix.columns[j] not in col_corr
                ):
                    colname = corr_matrix.columns[i]  # getting the name of column
                    col_corr.add(colname)
                    if colname in self.df.columns:
                        list1.append(colname)
                        del self.df[colname]  # deleting the column from the self.df
                    if colname in numerical_variables:
                        numerical_variables.remove(colname)
        print("List of columns removed due to high correlation: ", list1)
        return pd.DataFrame(self.df)

    def cramers_v_matrix(self, category_variables: list, threshold: float = 0.1):
        """
        Method to remove the variables with cramers v value greater than threshold
        :param category_variables: list - list of categorical variables to consider for Cramer's V analysis.
        :param threshold: float - threshold value for Cramer's V. Variables with Cramer's V greater than this threshold will be removed from the dataset.
        :return: dataframe - cramers v matrix
        """

        def cramers_V(variable_1: str, variable_2: str):
            """
              Method to calculate the Cramer's V statistic for categorical-categorical association.
              :param variable_1: string - categorical variable
              :param variable_2: string - categorical variable
              :return: float
              """
            crosstab = np.array(
                pd.crosstab(variable_1, variable_2, rownames=None, colnames=None))  # Cross table building
            stat = chi2_contingency(crosstab)[0]  # Keeping of the test statistic of the Chi2 test
            obs = np.sum(crosstab)  # Number of observations
            mini = min(
                crosstab.shape) - 1  # Take the minimum value between the columns and the rows of the cross table
            return stat / (obs * mini)

        if not category_variables:
            print("No category variables found")
            return "No category variables found"
        data_df_orig = self.df[category_variables]
        rows = []

        for var1 in data_df_orig:
            col = []
            for var2 in data_df_orig:
                cramers = cramers_V(data_df_orig[var1], data_df_orig[var2])  # Cramer's V test
                col.append(round(cramers, 2))  # Keeping of the rounded value of the Cramer's V
            rows.append(col)

        cramers_results = np.array(rows)
        df_matrix = pd.DataFrame(cramers_results, columns=data_df_orig.columns, index=data_df_orig.columns)

        list1 = list()
        col_cramers = set()  # Set of all the names of deleted columns

        for i in range(len(df_matrix.columns)):
            for j in range(i):
                if (df_matrix.iloc[i, j] >= threshold) and (df_matrix.columns[j] not in col_cramers):
                    colname = df_matrix.columns[i]  # getting the name of column
                    col_cramers.add(colname)
                    if colname in data_df_orig.columns:
                        list1.append(colname)
                        del data_df_orig[colname]  # deleting the column from the dataset
                        del self.df[colname]
                        category_variables.remove(colname)
        # self.df=data_df_orig
        print("List of columns removed due to higher than the threshold cramers v value: ", list1)
        df_matrix.reset_index(inplace=True)
        df_matrix.rename(columns={'index': 'variable'}, inplace=True)
        return df_matrix

    def outlier_capping(self, column_list: list, thold: float = 3):
        """
        Method to cap the outliers
        :param column_list: list - list of columns to consider for outlier capping
        :param thold: float - threshold value for outlier capping
        :return: pd.DataFrame - dataframe with capped outliers
        """
        for col in column_list:
            mu = self.df[col].mean()
            sigma = self.df[col].std()
            upper_limit = mu + (thold * sigma)
            lower_limit = mu - (thold * sigma)
            self.df[col] = np.clip(self.df[col], lower_limit, upper_limit)
        return self.df

    def outlier_percentages(self, column_list: list, thold: float = 3):
        """
        Method to calculate the percentage of outliers
        :param column_list: list - list of columns to consider for outlier capping
        :param thold: float - threshold value for outlier capping
        :return: pd.DataFrame - dataframe with outlier percentages
        """
        outlier_percentages = {}
        for col in column_list:
            mu = self.df[col].mean()
            sigma = self.df[col].std()
            upper_limit = mu + (thold * sigma)
            lower_limit = mu - (thold * sigma)
            outliers = self.df[(self.df[col] > upper_limit) | (self.df[col] < lower_limit)]
            outlier_percentages[col] = round(len(outliers) / len(self.df) * 100, 2)
        outlier_percentages = pd.DataFrame(outlier_percentages.items(),
                                           columns=['column', 'outlier_percentage'])
        outlier_percentages = outlier_percentages[outlier_percentages['outlier_percentage'] != 0]
        outlier_percentages.reset_index(drop=True, inplace=True)
        return outlier_percentages

    def standardization(self):
        """
        Method to standardize the data
        :return: pd.DataFrame - standardized dataframe
        """
        scaler = StandardScaler()
        scaled = scaler.fit_transform(self.df)
        self.df = pd.DataFrame(scaled, columns=self.df.columns)
        return self.df

    def save_preprocessed_data(self, filename: str):
        """
        Method to save the data to a csv file
        :param filename: str - path to the file
        :return:
        """
        self.df.to_csv(filename)


class UVyper:
    def __init__(self, preprocessed_dataset: str, outlier_per: pd.DataFrame = None, cramers_matrix: pd.DataFrame = None,
                 missing_values_table: pd.DataFrame = None):
        """
        Method to read and initialize the data.
        :param preprocessed_dataset: str - path to the preprocessed dataset
        :param outlier_per: pd.DataFrame - dataframe with outlier percentages (optional)
        :param cramers_matrix: pd.DataFrame - dataframe with cramers matrix (optional)
        :param missing_values_table: pd.DataFrame - dataframe with missing values table (optional)
        """
        self.df = pd.read_csv(preprocessed_dataset)
        self.score_table = pd.DataFrame()
        self.distribution = pd.DataFrame()
        self.outlier_per = outlier_per
        self.cramers_matrix = cramers_matrix
        self.missing = missing_values_table

    def kmeans_w(self, minK: int, maxK: int, metric: str, min_size_per: float, max_size_per: float,
                 rand_sample_prop: float, filename: str, dataset: str,
                 n_clusters: int = None, option: int = 0):
        """
        Method to find the clusters using KMeans.
        :param minK: int - The minimum number of clusters to consider.
        :param maxK: int - The maximum number of clusters to consider.
        :param metric: str - optional (default='distortion'). The metric used to quantify the
        quality of clustering. Possible options include distortion, silhouette, calinski_harabasz,
        davies_bouldin, and others.
        :param min_size_per: float - percentage of minimum size of cluster
        :param max_size_per: float - percentage of maximum size of cluster
        :param rand_sample_prop: float - random sampling proportion
        :param filename:  float - path of the pickle file
        :param dataset: float - path of the original dataset
        :param n_clusters: int - no of clusters
        :param option: int - 1 - to save the model and 0 - not to save the model
        :return: ndarray - cluster labels
        """

        def elbow(minK: int, maxK: int, metric: str = 'distortion'):
            """
            Method to identify the optimal number of clusters (k) for a given dataset using the elbow method.
            :param minK: int - The minimum number of clusters to consider.
            :param maxK: int - The maximum number of clusters
            to consider.
            :param metric: str - optional (default='distortion'). The metric used to quantify the
            quality of clustering. Possible options include distortion, silhouette, calinski_harabasz,
            davies_bouldin, and others.
            :return: int - The optimal number of clusters (k).
            """
            X = self.df
            model = KMeans()
            visualizer = KElbowVisualizer(model, k=(minK, maxK), metric=metric, timings=False)
            visualizer.fit(X)
            visualizer.show()
            return visualizer.elbow_value_

        def kmeans(n_clusters: int, min_size_per: float, max_size_per: float):
            """
            Method to find the clusters using KMeans.
            :param n_clusters: no of clusters
            :param min_size_per: percentage of minimum size of cluster
            :param max_size_per: percentage of maximum size of cluster
            :return: ndarray - The cluster labels.
            """
            kmeanModel = KMeansConstrained(
                n_clusters=n_clusters,
                size_min=int((min_size_per / 100) * self.df.shape[0]),
                size_max=int((max_size_per / 100) * self.df.shape[0]), )
            clusters = kmeanModel.fit_predict(self.df)
            return clusters

        def kmeans_model_create(n_clusters: int, min_size_per: float, max_size_per: float, filename: str,
                                rand_sample_prop: float = 1.0):
            """
            Method to create the KMeans model as a pickle file.
            :param filename: name of the pickle file
            :param n_clusters: no of clusters
            :param min_size_per: percentage of minimum size of cluster
            :param max_size_per: percentage of maximum size of cluster
            :param rand_sample_prop: random sampling proportion
            """
            sample_data = self.df.sample(frac=rand_sample_prop, random_state=42)
            kmeanModel = KMeansConstrained(
                n_clusters=n_clusters,
                size_min=int((min_size_per / 100) * sample_data.shape[0]),
                size_max=int((max_size_per / 100) * sample_data.shape[0]),
            )
            kmeanModel.fit(sample_data)
            with open(filename, 'wb') as file:
                pickle.dump(kmeanModel, file)

        def kmeans_model_read(filename: str):
            """
            Method to read the KMeans model from a pickle file.
            :param filename: str - The path of the pickle file.
            :return: ndarray - The cluster labels.
            """
            with open(filename, 'rb') as file:
                kmeanModel = pickle.load(file)
            clusters = kmeanModel.predict(self.df)
            return clusters

        def get_cluster_centers(clusters: np.ndarray, dataset: str):
            """
            Method to calculate the cluster centers.
            :param dataset: str - path to the dataset
            :param clusters: list - The cluster labels.
            :return: DataFrame - The cluster centers.
            """
            temp = pd.read_csv(dataset)
            temp['cluster'] = clusters
            return temp.groupby(clusters).mean()

        def get_cluster_distribution(clusters: np.ndarray):
            """
            Method to calculate the distribution of clusters.
            :param clusters: ndarray - The cluster labels.
            :return: dataframe - The distribution of clusters.
            """
            df = pd.DataFrame()
            df['cluster'] = sorted(set(clusters))
            df['count'] = [list(clusters).count(i) for i in sorted(set(clusters))]
            df['percentage'] = round(df['count'] / df['count'].sum(), 2)
            df['Model'] = ['Kmeans' for _ in range(len(set(clusters)))]
            df = df.reset_index(drop=True)
            return df

        def get_scores(clusters: np.ndarray):
            """
            Method to calculate the silhouette score and Davies-Bouldin score.
            :param clusters: list - The cluster labels.
            :return: float - The silhouette score.
            :return: float - The Davies-Bouldin score.
            """
            scores = pd.DataFrame()
            sil = silhouette_score(self.df, clusters)
            dav = davies_bouldin_score(self.df, clusters)
            scores['Model'] = ['KMeans']
            scores['Silhouette'] = [sil]
            scores['Davies Bouldin'] = [dav]
            scores['n_clusters'] = [len(set(clusters))]
            return scores

        print("Performing KMeans Clustering...")
        if n_clusters is None:
            n_clusters = elbow(minK, maxK, metric)
        if option == 1:
            kmeans_model_create(n_clusters=n_clusters, min_size_per=min_size_per, max_size_per=max_size_per,
                                filename=filename, rand_sample_prop=rand_sample_prop)
            clusters = kmeans_model_read(filename)
        else:
            clusters = kmeans(n_clusters=n_clusters, min_size_per=min_size_per, max_size_per=max_size_per)
        cluster_centers = get_cluster_centers(clusters, dataset)
        cluster_distribution = get_cluster_distribution(clusters)
        scores = get_scores(clusters)
        self.score_table = self.score_table.append(scores, ignore_index=True)
        self.distribution = self.distribution.append(cluster_distribution, ignore_index=True)
        print(cluster_distribution)
        print(scores)
        print(cluster_centers)
        print("KMeans Clustering Complete!")
        return clusters

    def hierarchical_w(self, param_grid: dict, folds: int, n_iter: int, rand_sample_prop: float, dataset: str,
                       n_clusters: int = None, linkage: str = None,
                       affinity: str = None):
        """
        Method to perform hierarchical clustering.
        :param param_grid: dict - parameters grid
        :param folds: int - number of folds
        :param n_iter: int - number of iterations
        :param rand_sample_prop: float - random sample proportion
        :param dataset: str - path to the original dataset
        :param n_clusters: int - number of clusters
        :param linkage: str - linkage
        :param affinity: str - affinity
        :return: ndarray - cluster labels
        """

        def silhouettee(estimator: object, df: pd.DataFrame, metric: str = 'euclidean'):
            """
            Method to calculate the silhouette score
            :param estimator: object - estimator
            :param df: dataframe - data
            :param metric: str - metric
            :return: float - silhouette score
            """
            labels = estimator.fit_predict(df)
            score = silhouette_score(df, labels, metric=metric)
            return score

        def randomizedSearchCV_hierarchical(grid: dict, cv: int = 5, n_iter: int = 10, rand_sample_prop: float = 0.2):
            """
            Method to perform randomized search cross validation
            :param grid: dict - grid
            :param cv: int - number of folds
            :param n_iter: int - number of iterations
            :param rand_sample_prop: float - random sample proportion
            :return: tuple - best score, best parameters, training history
            """
            sample_data = self.df.sample(frac=rand_sample_prop)
            hierarchical = AgglomerativeClustering()
            random_search = RandomizedSearchCV(estimator=hierarchical, param_distributions=grid, cv=cv,
                                               scoring=silhouettee, n_iter=n_iter)
            random_search.fit(sample_data)
            training_history = random_search.cv_results_
            return random_search.best_score_, random_search.best_params_, training_history

        def hierarchical(n_clusters: int, linkage: str, affinity: str, random_sample_prop: float = 1.0):
            """
            Method to perform hierarchical clustering
            :param n_clusters: int - number of clusters
            :param linkage: str - linkage
            :param affinity: str - affinity
            :param random_sample_prop: float - random sample proportion
            :return: ndarray - clusters
            """
            sample_data = self.df.sample(frac=random_sample_prop, random_state=42)
            hierarchical = AgglomerativeClustering(n_clusters=n_clusters, affinity=affinity, linkage=linkage)
            clusters = hierarchical.fit_predict(sample_data)
            return clusters, sample_data

        def knn(sample_data: pd.DataFrame, clusters: np.ndarray, n_neighbors: int = 5):
            """
            Method to perform knn
            :param sample_data: dataframe - sample data used in hierarchical method
            :param clusters: ndarray - clusters from hierarchical method
            :param n_neighbors: int - number of neighbors
            :return: ndarray - clusters labels for the entire dataset
            """
            knn = KNeighborsClassifier(n_neighbors=n_neighbors)
            knn.fit(sample_data, clusters)
            clu = knn.predict(self.df)
            return clu

        # @staticmethod
        def get_cluster_centers(clusters: np.ndarray, dataset: str):
            """
            Method to calculate the cluster centers.
            :param dataset: str - path to the dataset
            :param clusters: list - The cluster labels.
            :return: DataFrame - The cluster centers.
            """
            temp = pd.read_csv(dataset)
            temp['cluster'] = clusters
            return temp.groupby(clusters).mean()

        def get_cluster_distribution(clusters: np.ndarray):
            """
            Method to calculate the distribution of clusters.
            :param clusters: ndarray - The cluster labels.
            :return: dataframe - The distribution of clusters.
            """
            df = pd.DataFrame()
            df['cluster'] = sorted(set(clusters))
            df['count'] = [list(clusters).count(i) for i in sorted(set(clusters))]
            df['percentage'] = round(df['count'] / df['count'].sum(), 2)
            df['Model'] = ['Hierarchical' for _ in range(len(set(clusters)))]
            df = df.reset_index(drop=True)
            return df

        def get_scores(clusters: np.ndarray):
            """
            Method to calculate the silhouette score and Davies-Bouldin score.
            :param clusters: list - The cluster labels.
            :return: float - The silhouette score.
            :return: float - The Davies-Bouldin score.
            """
            scores = pd.DataFrame()
            sil = silhouette_score(self.df, clusters)
            dav = davies_bouldin_score(self.df, clusters)
            scores['Model'] = ['Hierarchical']
            scores['Silhouette'] = [sil]
            scores['Davies Bouldin'] = [dav]
            scores['n_clusters'] = [len(set(clusters))]
            return scores

        print("Performing Hierarchical Clustering...")
        if n_clusters is None or linkage is None or affinity is None:
            a, b, c = randomizedSearchCV_hierarchical(grid=param_grid, cv=folds, n_iter=n_iter,
                                                      rand_sample_prop=rand_sample_prop)
            if n_clusters is None:
                n_clusters = b['n_clusters']
                print("Recommended number of clusters: ", n_clusters)
            if linkage is None:
                linkage = b['linkage']
                print("Recommended linkage: ", linkage)
            if affinity is None:
                affinity = b['affinity']
                print("Recommended affinity: ", affinity)

        clusters, sample_data = hierarchical(n_clusters=n_clusters, linkage=linkage, affinity=affinity,
                                             random_sample_prop=rand_sample_prop)
        clusters = knn(sample_data, clusters)
        cluster_centers = get_cluster_centers(clusters, dataset)
        cluster_distribution = get_cluster_distribution(clusters)
        scores = get_scores(clusters)
        self.distribution = self.distribution.append(cluster_distribution, ignore_index=True)
        self.score_table = self.score_table.append(scores, ignore_index=True)
        print(cluster_distribution)
        print(scores)
        print(cluster_centers)
        print("Hierarchical Clustering Complete!")
        return clusters

    def gmm_w(self, param_grid: dict, folds: int, n_iter: int, rand_sample_prop: float, filename: str, dataset: str,
              n_components: int = None,
              covariance_type: str = None,
              init_params: str = None, option: int = 0):
        """
        Method to perform Gaussian Mixture Model clustering.
        :param param_grid: dict - The parameters to be used for the randomized search cross validation.
        :param folds: int - The number of folds to be used for the cross validation.
        :param n_iter: int - The number of iterations to be used for the randomized search cross validation.
        :param rand_sample_prop: float - random sample proportion
        :param filename: str - path to the pickle file
        :param dataset: str - path to the original dataset
        :param n_components: int - number of components
        :param covariance_type: str - covariance type
        :param init_params: str - initialization parameters
        :param option: int - 1 - to save the model and 0 - not to save the model
        :return: ndarray - The cluster labels.
        """

        def silhouettee(estimator: object, df: pd.DataFrame, metric: str = 'euclidean'):
            """
            Method to calculate the silhouette score.
            :param estimator: object - clustering algorithm that implements fit_predict method
            :param df: DataFrame - data
            :param metric: str - metric to use for calculating the distance between points
            :return: float - silhouette score
            """
            labels = estimator.fit_predict(df)
            score = silhouette_score(df, labels, metric=metric)
            return score

        def randomizedSearchCV_gmm(grid: dict, cv: int = 5, n_iter: int = 10, rand_sample_prop: float = 0.2):
            """
            Method to perform randomized search cross validation on the GMM algorithm.
            :param grid: dict - dictionary of parameters to search over
            :param cv: int - number of folds
            :param n_iter: int - number of iterations
            :param rand_sample_prop: float - proportion of data to sample
            :return: tuple - best score, the best parameters, and cross validation results
            """
            sample_data = self.df.sample(frac=rand_sample_prop)
            gmm = GaussianMixture()
            random_search = RandomizedSearchCV(estimator=gmm, param_distributions=grid, cv=cv, scoring=silhouettee,
                                               n_iter=n_iter)
            random_search.fit(sample_data)
            return random_search.best_score_, random_search.best_params_, random_search.cv_results_

        def gmm(n_components: int, covariance_type: str, init_params: str):
            """
            Method to perform GMM clustering.
            :param n_components: int - number of clusters
            :param covariance_type: str - covariance type
            :param init_params: str - initialization method
            :return: numpy array - cluster labels
            """
            gmm = GaussianMixture(n_components=n_components, covariance_type=covariance_type, init_params=init_params)
            clusters = gmm.fit_predict(self.df)
            return clusters

        def gmm_model_create(n_components: int, covariance_type: str, init_params: str, filename: str,
                             rand_sample_prop: float = 1.0):
            """
            Method to create pickle file of gmm model.
            :param filename: name of the pickle file
            :param n_components: int - number of clusters
            :param covariance_type: str - covariance type
            :param init_params: str - initialization method
            :param rand_sample_prop: float - proportion of data to sample
            :return: numpy array - cluster labels
            """
            sample_data = self.df.sample(frac=rand_sample_prop, random_state=42)
            gmm = GaussianMixture(n_components=n_components, covariance_type=covariance_type, init_params=init_params)
            gmm.fit(sample_data)
            with open(filename, 'wb') as f:
                pickle.dump(gmm, f)

        def gmm_model_read(filename: str):
            """
            Method to load the pickle file of gmm model.
            :param filename: str - path to the pickle file
            :return: ndarray - cluster labels
            """
            with open(filename, 'rb') as f:
                gmm = pickle.load(f)
            clusters = gmm.predict(self.df)
            return clusters

        def get_cluster_centers(clusters: np.ndarray, dataset: str):
            """
            Method to calculate the cluster centers.
            :param dataset: str - path to the dataset
            :param clusters: list - The cluster labels.
            :return: DataFrame - The cluster centers.
            """
            temp = pd.read_csv(dataset)
            temp['cluster'] = clusters
            return temp.groupby(clusters).mean()

        def get_cluster_distribution(clusters: np.ndarray):
            """
            Method to calculate the distribution of clusters.
            :param clusters: ndarray - The cluster labels.
            :return: dataframe - The distribution of clusters.
            """
            df = pd.DataFrame()
            df['cluster'] = sorted(set(clusters))
            df['count'] = [list(clusters).count(i) for i in sorted(set(clusters))]
            df['percentage'] = round(df['count'] / df['count'].sum(), 2)
            df['Model'] = ['GMM' for _ in range(len(set(clusters)))]
            df = df.reset_index(drop=True)
            return df

        def get_scores(clusters: np.ndarray):
            """
            Method to calculate the silhouette score and Davies-Bouldin score.
            :param clusters: list - The cluster labels.
            :return: float - The silhouette score.
            :return: float - The Davies-Bouldin score.
            """
            scores = pd.DataFrame()
            if len(set(clusters)) == 1:
                sil = 0
                dav = 0
            else:
                sil = silhouette_score(self.df, clusters)
                dav = davies_bouldin_score(self.df, clusters)
            scores['Model'] = ['GMM']
            scores['Silhouette'] = [sil]
            scores['Davies Bouldin'] = [dav]
            scores['n_clusters'] = [len(set(clusters))]
            return scores

        print("Performing GMM Clustering...")
        if n_components is None or covariance_type is None or init_params is None:
            a, b, c = randomizedSearchCV_gmm(grid=param_grid, cv=folds, n_iter=n_iter,
                                             rand_sample_prop=rand_sample_prop)
            if n_components is None:
                n_components = b['n_components']
                print("Recommended number of clusters: ", n_components)
            if covariance_type is None:
                covariance_type = b['covariance_type']
                print("Recommended covariance type: ", covariance_type)
            if init_params is None:
                init_params = b['init_params']
                print("Recommended initialization method: ", init_params)
        if option == 1:
            gmm_model_create(n_components=n_components, covariance_type=covariance_type, init_params=init_params,
                             filename=filename, rand_sample_prop=rand_sample_prop)
            clusters = gmm_model_read(filename=filename)
        else:
            clusters = gmm(n_components=n_components, covariance_type=covariance_type, init_params=init_params)
        cluster_centers = get_cluster_centers(clusters, dataset)
        cluster_distribution = get_cluster_distribution(clusters)
        scores = get_scores(clusters)
        self.distribution = self.distribution.append(cluster_distribution, ignore_index=True)
        self.score_table = self.score_table.append(scores, ignore_index=True)
        print(cluster_distribution)
        print(scores)
        print(cluster_centers)
        print("GMM Clustering Complete!")
        return clusters

    def birch_w(self, param_grid: dict, folds: int, n_iter: int, rand_sample_prop: float, filename: str, dataset: str,
                n_clusters: int = None,
                branching_factor: int = None,
                threshold: float = None, option: int = 0):
        """
        Method to perform Birch clustering.
        :param param_grid: dict - The parameters to be used for the randomized search cross validation.
        :param folds: int - The number of folds to be used for the cross validation.
        :param n_iter: int - The number of iterations to be used for the randomized search cross validation.
        :param rand_sample_prop: float - random sample proportion
        :param filename: str - path to the pickle file
        :param dataset: str - path to the original dataset
        :param n_clusters: int - number of clusters
        :param branching_factor: int - branching factor
        :param threshold: int - threshold
        :param option: int - 1 - to save the model and 0 - not to save the model
        :return: ndarray - The cluster labels.
        """

        def silhouettee(estimator: object, df: pd.DataFrame, metric: str = 'euclidean'):
            """
            Method to calculate the silhouette score.
            :param estimator: object - clustering algorithm that implements fit_predict method
            :param df: DataFrame - data
            :param metric: str - metric to use for calculating the distance between points
            :return: float - silhouette score
            """
            labels = estimator.fit_predict(df)
            score = silhouette_score(df, labels, metric=metric)
            return score

        def randomizedSearchCV_birch(grid: dict, cv: int = 5, n_iter: int = 10, rand_sample_prop: float = 0.2):
            """
            Method to perform randomized search cross validation on the GMM algorithm.
            :param grid: dict - dictionary of parameters to search over
            :param cv: int - number of folds
            :param n_iter: int - number of iterations
            :param rand_sample_prop: float - proportion of data to sample
            :return: tuple - best score, the best parameters, and cross validation results
            """
            sample_data = self.df.sample(frac=rand_sample_prop)
            birch = Birch()
            random_search = RandomizedSearchCV(estimator=birch, param_distributions=grid, cv=cv, scoring=silhouettee,
                                               n_iter=n_iter)
            random_search.fit(sample_data)
            return random_search.best_score_, random_search.best_params_, random_search.cv_results_

        def birch(n_clusters: int, branching_factor: int, threshold: float):
            """
            Method to perform GMM clustering.
            :param n_clusters: int - number of clusters
            :param branching_factor: int - branching factor
            :param threshold: int - threshold
            :return: numpy array - cluster labels
            """
            birch = Birch(n_clusters=n_clusters, branching_factor=branching_factor, threshold=threshold)
            clusters = birch.fit_predict(self.df)
            return clusters

        def birch_model_create(n_clusters: int, branching_factor: int, threshold: float, filename: str,
                               rand_sample_prop: float = 1.0):
            """
            Method to create pickle file of gmm model.
            :param filename: str - path to the pickle file
            :param n_clusters: int - number of clusters
            :param branching_factor: int - branching factor
            :param threshold: int - threshold
            :param rand_sample_prop: float - proportion of data to sample
            :return: numpy array - cluster labels
            """
            sample_data = self.df.sample(frac=rand_sample_prop, random_state=42)
            birch = Birch(n_clusters=n_clusters, branching_factor=branching_factor, threshold=threshold)
            birch.fit(sample_data)
            with open(filename, 'wb') as f:
                pickle.dump(birch, f)

        def birch_model_read(filename: str):
            """
            Method to load the pickle file of gmm model.
            :param filename: str - path to the pickle file
            :return: ndarray - cluster labels
            """
            with open(filename, 'rb') as f:
                birch = pickle.load(f)
            clusters = birch.predict(self.df)
            return clusters

        def get_cluster_centers(clusters: np.ndarray, dataset: str):
            """
            Method to calculate the cluster centers.
            :param dataset: str - path to the dataset
            :param clusters: list - The cluster labels.
            :return: DataFrame - The cluster centers.
            """
            temp = pd.read_csv(dataset)
            temp['cluster'] = clusters
            return temp.groupby(clusters).mean()

        def get_cluster_distribution(clusters: np.ndarray):
            """
            Method to calculate the distribution of clusters.
            :param clusters: ndarray - The cluster labels.
            :return: dataframe - The distribution of clusters.
            """
            df = pd.DataFrame()
            df['cluster'] = sorted(set(clusters))
            df['count'] = [list(clusters).count(i) for i in sorted(set(clusters))]
            df['percentage'] = round(df['count'] / df['count'].sum(), 2)
            df['Model'] = ['Birch' for _ in range(len(set(clusters)))]
            df = df.reset_index(drop=True)
            return df

        def get_scores(clusters: np.ndarray):
            """
            Method to calculate the silhouette score and Davies-Bouldin score.
            :param clusters: list - The cluster labels.
            :return: float - The silhouette score.
            :return: float - The Davies-Bouldin score.
            """
            scores = pd.DataFrame()
            sil = silhouette_score(self.df, clusters)
            dav = davies_bouldin_score(self.df, clusters)
            scores['Model'] = ['Birch']
            scores['Silhouette'] = [sil]
            scores['Davies Bouldin'] = [dav]
            scores['n_clusters'] = [len(set(clusters))]
            return scores

        print("Performing Birch Clustering...")
        if n_clusters is None or branching_factor is None or threshold is None:
            a, b, c = randomizedSearchCV_birch(grid=param_grid, cv=folds, n_iter=n_iter,
                                               rand_sample_prop=rand_sample_prop)
            if n_clusters is None:
                n_clusters = b['n_clusters']
                print("Recommended number of clusters: ", n_clusters)
            if branching_factor is None:
                branching_factor = b['branching_factor']
                print("Recommended branching factor: ", branching_factor)
            if threshold is None:
                threshold = b['threshold']
                print("Recommended threshold: ", threshold)

        if option == 1:
            birch_model_create(n_clusters=n_clusters, branching_factor=branching_factor, threshold=threshold,
                               filename=filename, rand_sample_prop=rand_sample_prop)
            clusters = birch_model_read(filename=filename)
        else:
            clusters = birch(n_clusters=n_clusters, branching_factor=branching_factor, threshold=threshold)
        cluster_centers = get_cluster_centers(clusters, dataset)
        cluster_distribution = get_cluster_distribution(clusters)
        scores = get_scores(clusters)
        self.distribution = self.distribution.append(cluster_distribution, ignore_index=True)
        self.score_table = self.score_table.append(scores, ignore_index=True)
        print(cluster_distribution)
        print(scores)
        print(cluster_centers)
        print("Birch Clustering Complete!")
        return clusters

    def get_models_summary(self):
        """
        Method to get the summary of the clustering
        prints the score table and best clustering model
        :return: str - recommended model
        """

        def sort_score_table(score_df: pd.DataFrame):
            silhouette_ranks = score_df['Silhouette'].rank(method='dense', ascending=False)
            db_ranks = score_df['Davies Bouldin'].rank(method='dense')
            combined_ranks = (silhouette_ranks + db_ranks) / 2
            score_df['Rank'] = combined_ranks.rank(method='dense')
            score_df = score_df.sort_values(by=['Rank'])
            return score_df

        ranked_score_table = sort_score_table(self.score_table)
        self.score_table = ranked_score_table
        print(ranked_score_table)
        recommended_model = ranked_score_table.iloc[0]['Model']
        print("Recommended Model: ", recommended_model)
        return recommended_model

    @staticmethod
    def post_process(recommended_model: str, org_dataset: str, preprocessed_dataset: str, dependent_variable: str,
                     filename: str, kmeans_cluster_labels: np.ndarray = None,
                     hierarchical_cluster_labels: np.ndarray = None,
                     gmm_cluster_labels: np.ndarray = None,
                     birch_cluster_labels: np.ndarray = None, n_variables: int = 5, ):

        """
        Method to perform post-processing of the clustering results
        :param recommended_model: str - name of the recommended model
        :param org_dataset: str - path to the original dataset
        :param preprocessed_dataset: str - path to the preprocessed dataset
        :param dependent_variable: str - name of the dependent variable
        :param filename: str - filename to save the parallel coordinates plot (with .png extension)
        :param kmeans_cluster_labels: np.ndarray - cluster labels of KMeans (optional)
        :param hierarchical_cluster_labels: np.ndarray - cluster labels of Hierarchical (optional)
        :param gmm_cluster_labels: np.ndarray - cluster labels of GMM (optional)
        :param birch_cluster_labels: np.ndarray - cluster labels of Birch (optional)
        :param n_variables: int - number of differential factors to be considered
        :return: pd.DataFrame - reduced preprocessed dataset to 2 dimensions using PCA
        """

        def save_clustered_dataset(recommended_model: str, org_dataset: str,
                                   kmeans_cluster_labels: np.ndarray, hierarchical_cluster_labels: np.ndarray,
                                   gmm_cluster_labels: np.ndarray,
                                   birch_cluster_labels: np.ndarray):

            temp = pd.read_csv(org_dataset)
            if recommended_model == 'KMeans':
                temp['cluster'] = kmeans_cluster_labels
                # temp.to_csv('KMeans_Clustered_' + org_dataset, index=False)
            elif recommended_model == 'Hierarchical':
                temp['cluster'] = hierarchical_cluster_labels
                # temp.to_csv('Hierarchical_Clustered_' + org_dataset, index=False)
            elif recommended_model == 'GMM':
                temp['cluster'] = gmm_cluster_labels
                # temp.to_csv('GMM_Clustered_' + org_dataset, index=False)
            elif recommended_model == 'Birch':
                temp['cluster'] = birch_cluster_labels
                # temp.to_csv('Birch_Clustered_' + org_dataset, index=False)

            return temp

        def impute_na(df, columns_list: list, method: str):  # mean, mode,  bfill, ffill
            if method == "mean":
                for i in columns_list:
                    if i in df.columns:
                        df[i].fillna(df[i].mean(), inplace=True)
            if method == "mode":
                for i in columns_list:
                    if i in df.columns:
                        df[i].fillna(df[i].mode()[0], inplace=True)

            if method == "ffill" or method == "bfill":
                for i in columns_list:
                    if i in df.columns:
                        df[i].fillna(method=method, inplace=True)

        def parallel_plot(df: pd.DataFrame, features: list, filename: str):
            """
            Method to plot parallel coordinates
            :param filename: str - filename to save the plot (with .png extension)
            :param df: pd.DataFrame - dataframe to be plotted
            :param features: list - list of features to be plotted
            :return:
            """
            fig = px.parallel_coordinates(df, color='cluster', dimensions=features,
                                          color_continuous_scale=px.colors.diverging.Tealrose, )
            fig.show()
            fig.update_layout(width=1904)
            fig.update_layout(height=959)
            pio.write_image(fig, filename)

        def pca(preprocessed_data: pd.DataFrame, clusters: np.ndarray, n_components: int = 2):
            """
            Method to perform PCA.
            :param preprocessed_data: dataframe - The dataset.
            :param clusters: ndarray - The cluster labels.
            :param n_components: int - The number of components to consider.
            :return: dataframe - The PCA results.
            """
            pca = PCA(n_components=n_components)
            principalComponents = pca.fit_transform(preprocessed_data)
            principalDf = pd.DataFrame(data=principalComponents,
                                       columns=['PC' + str(i) for i in range(1, n_components + 1)])
            principalDf['cluster'] = clusters
            return principalDf

        def scatter_plot_2d(principalComponents: pd.DataFrame, recommended_model: str):
            """
            Method to plot the PCA results.
            :param principalComponents: dataframe - The PCA results.
            :param recommended_model: str - name of the recommended model
            :return: Scatter Plot
            """
            plt.scatter(principalComponents['PC1'], principalComponents['PC2'], c=principalComponents['cluster'],
                        cmap=viridis)
            plt.title(recommended_model + ' clustering')
            plt.colorbar()
            plt.show()

        def min_max_scaling(df: pd.DataFrame):
            """
            Method to perform min max scaling.
            :param df: dataframe - The dataset.
            :return: dataframe - The scaled dataset.
            """
            scaler = MinMaxScaler()
            df = pd.DataFrame(scaler.fit_transform(df), columns=df.columns)
            return df

        def randomforest(df, cluster_label):
            cluster_df = pd.DataFrame()
            cluster_df['cluster'] = df['cluster']
            df.drop('cluster', axis=1, inplace=True)
            cluster_df['Binary Cluster ' + str(cluster_label)] = cluster_df['cluster'].apply(
                lambda x: 1 if x == cluster_label else 0)
            print("Cluster " + str(cluster_label) + " classification counts:")
            print("\n", cluster_df["Binary Cluster " + str(cluster_label)].value_counts())
            # Train a classifier
            clf = RandomForestClassifier(random_state=1)
            clf.fit(df[df.columns].values, cluster_df["Binary Cluster " + str(cluster_label)].values)
            # Index sort the most important features
            sorted_feature_weight_idxes = np.argsort(clf.feature_importances_)[::-1]  # Reverse sort

            # Get the most important features names and weights
            most_important_features = np.take_along_axis(np.array(df.columns.tolist()), sorted_feature_weight_idxes,
                                                         axis=0)
            most_important_weights = np.take_along_axis(np.array(clf.feature_importances_), sorted_feature_weight_idxes,
                                                        axis=0)
            ranks = np.arange(1, len(most_important_weights) + 1)
            df['cluster'] = cluster_df['cluster']
            return most_important_features, ranks

        clustered_dataset = save_clustered_dataset(recommended_model=recommended_model, org_dataset=org_dataset,
                                                   kmeans_cluster_labels=kmeans_cluster_labels,
                                                   hierarchical_cluster_labels=hierarchical_cluster_labels,
                                                   gmm_cluster_labels=gmm_cluster_labels,
                                                   birch_cluster_labels=birch_cluster_labels)
        preprocessed_data = pd.read_csv(preprocessed_dataset)
        principalComponents = pca(preprocessed_data, clustered_dataset['cluster'], n_components=2)
        temp = principalComponents.drop('cluster', axis=1)
        scatter_plot_2d(principalComponents, recommended_model)
        df = clustered_dataset.copy()
        df = df.select_dtypes(include=['float64', 'int64'])
        if dependent_variable in df.columns:
            df.drop(dependent_variable, axis=1, inplace=True)
        impute_na(df, df.select_dtypes(include=['float64', 'int64']).columns, 'mean')
        cluster_labels = df['cluster'].unique()
        feature_score_table = pd.DataFrame(columns=['Feature', 'Rank'])
        for cluster_label in cluster_labels:
            fea, ranks = randomforest(df, cluster_label)
            feature_score_table = feature_score_table.append(pd.DataFrame({'Feature': fea, 'Rank': ranks}),
                                                             ignore_index=True)
        feature_score_table = (feature_score_table.groupby(['Feature']).sum().sort_values(by='Rank',
                                                                                          ascending=True)) / len(
            cluster_labels)
        feature_score_table.reset_index(inplace=True)
        if n_variables > len(feature_score_table):
            n_variables = len(feature_score_table)
        features = list(feature_score_table['Feature'][:n_variables])
        print("Differential Factors: ")
        for feature in features:
            print(feature)
        differential_factors_df = clustered_dataset[features]
        differential_factors_df = min_max_scaling(differential_factors_df)
        differential_factors_df['cluster'] = clustered_dataset['cluster']
        parallel_plot(differential_factors_df, features, filename=filename)
        return temp

    def playbook(self, filename: str, org_dataset: str, dependent_variable: str, pca: pd.DataFrame, im: str,
                 to_delete: bool = 0,
                 kmeans_cluster_labels: str = None,
                 hierarchical_cluster_labels: str = None, gmm_cluster_labels: str = None,
                 birch_cluster_labels: str = None, ):
        """
        Method to generate playbook
        :param filename: str - filename to save the playbook
        :param org_dataset: str - path to the original dataset
        :param dependent_variable: str - dependent variable
        :param pca: pd.DataFrame - pca dataframe of preprocess dataset
        :param im: str - path to the parallel plot png file (with .png extension
        :param to_delete: int - to delete the parallel plot png file after generating playbook
        :param kmeans_cluster_labels: np.ndarray - kmeans cluster labels
        :param hierarchical_cluster_labels: np.ndarray - hierarchical cluster labels
        :param gmm_cluster_labels: np.ndarray - gmm cluster labels
        :param birch_cluster_labels: np.ndarray - birch cluster labels
        :return:
        """

        def distribution(workbook, distribution_table: pd.DataFrame, score_table: pd.DataFrame):
            df = distribution_table.pivot(index='Model', columns='cluster', values='percentage')
            df.reset_index(inplace=True)
            df.fillna(0, inplace=True)

            # Create new Excel workbook and worksheet
            ws = workbook.create_sheet(0)
            ws.title = 'Distribution'

            for i in range(1, df.shape[0] + 3):
                ws.row_dimensions[i].height = 25
            for i in range(1, df.shape[1] + 1):
                ws.column_dimensions[get_column_letter(i)].width = 25
            for i in range(1, len(df.columns) + 1):
                ws.cell(row=2, column=i).font = Font(bold=True)
                ws.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=i).fill = PatternFill("solid", fgColor="A9C4FE")
                ws.cell(row=2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Write data to worksheet
            ws.merge_cells('A1:' + get_column_letter(df.shape[1]) + '1')
            ws.cell(row=1, column=1).value = 'Distribution of Clusters'
            ws.cell(row=1, column=1).font = Font(bold=True, size=16, underline='single')
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            rows = dataframe_to_rows(df, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    # if type(value) != str and type(value) != int:
                    #     # value = str(round(float(value), 2)) + '%'
                    #     value = '{:.2%}'.format(value/100)
                    ws.cell(row=r_idx + 1, column=c_idx, value=value)
                    ws.cell(row=r_idx + 1, column=c_idx).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=r_idx + 1, column=c_idx).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            df2 = score_table
            for i in range(2, df2.shape[0] + 4):
                ws.row_dimensions[i + df.shape[0] + 1].height = 25
            for i in range(1, df2.shape[1] + 1):
                ws.column_dimensions[get_column_letter(i + df.shape[0] + 1)].width = 25
            for i in range(1, len(df2.columns) + 1):
                ws.cell(row=2 + df.shape[0] + 2, column=i).font = Font(bold=True)
                ws.cell(row=2 + df.shape[0] + 2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2 + df.shape[0] + 2, column=i).fill = PatternFill("solid", fgColor="A9C4FE")
                ws.cell(row=2 + df.shape[0] + 2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            ws.merge_cells(
                'A' + str(1 + df.shape[0] + 2) + ':' + get_column_letter(df2.shape[1]) + str(1 + df.shape[0] + 2))
            ws.cell(row=1 + df.shape[0] + 2, column=1).value = 'Score of Clusters'
            ws.cell(row=1 + df.shape[0] + 2, column=1).font = Font(bold=True, size=16, underline='single')
            ws.cell(row=1 + df.shape[0] + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            rows = dataframe_to_rows(df2, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if type(value) != str and type(value) != int:
                        value = round(value, 2)
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx, value=value)
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx).border = Border(top=thin, left=thin,
                                                                                           right=thin,
                                                                                           bottom=thin)
            # Create chart
            chart = BarChart()
            chart.type = 'col'
            chart.style = 10
            chart.title = 'Cluster Distribution'
            chart.y_axis.title = '%'
            chart.x_axis.title = 'Model'

            # Set chart data
            data = Reference(ws, min_col=2, min_row=2, max_col=df.shape[1], max_row=df.shape[0] + 2)
            categories = Reference(ws, min_col=1, min_row=3, max_row=df.shape[0] + 2)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Add chart to worksheet
            ws.add_chart(chart, 'G1')
            ws.sheet_view.showGridLines = False

        def summary(workbook, org_dataset: str, dependent_variable: str):
            df = pd.read_csv(org_dataset)
            if df[dependent_variable].dtype != 'int64' or df[dependent_variable].dtype != 'float64':
                le = LabelEncoder()
                df[dependent_variable] = le.fit_transform(df[dependent_variable])
            m = Model(data=df, dependent_variable=dependent_variable, na_drop_threshold=0.5, training_percentage=0.7,
                      model_type='linear')
            variables_list = list(m.variables.get_active_variables())
            variables_list = variables_list + [dependent_variable]
            var_prof_df = DataProfiler(df[variables_list], dependent_variable=dependent_variable)
            var_prof_df.create_var_profiling_ws(wb=workbook, sheet_name='Summary', sort_by_variance='desc')

        def describe(workbook, org_dataset: str):
            df = pd.read_csv(org_dataset)
            df = df.describe().T
            ws = workbook.create_sheet()
            ws.title = 'Describe'
            ws.merge_cells(
                'A1:' + get_column_letter(df.shape[1] + 1) + '1')
            ws.cell(row=1, column=1).value = 'Distribution of Clusters'
            ws.cell(row=1, column=1).font = Font(bold=True, size=16, underline='single')
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
            ws.delete_rows(idx=3)
            ws['A2'] = 'Feature'
            for i in range(1, df.shape[0] + 2):
                ws.row_dimensions[i + 1].height = 25
            for i in range(1, df.shape[1] + 2):
                ws.column_dimensions[get_column_letter(i)].width = 25
            for i in range(1, len(df.columns) + 2):
                ws.cell(row=2, column=i).font = Font(bold=True)
                ws.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=i).fill = PatternFill("solid", fgColor="A9C4FE")
                ws.cell(row=2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for i in range(1, len(df.columns) + 2):
                for j in range(1, len(df.index) + 2):
                    ws.cell(row=j + 1, column=i).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=j + 1, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.sheet_view.showGridLines = False

        def analysis(workbook, outlier_per: pd.DataFrame = None, missing_table: pd.DataFrame = None,
                     cramers_table: pd.DataFrame = None):
            df = missing_table
            ws = workbook.create_sheet()
            ws.title = 'Analysis'
            for i in range(1, df.shape[0] + 3):
                ws.row_dimensions[i].height = 25
            for i in range(1, df.shape[1] + 1):
                ws.column_dimensions[get_column_letter(i)].width = 25
            for i in range(1, len(df.columns) + 1):
                ws.cell(row=2, column=i).font = Font(bold=True)
                ws.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=i).fill = PatternFill("solid", fgColor="A9C4FE")
                ws.cell(row=2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Write data to worksheet
            ws.merge_cells('A1:' + get_column_letter(df.shape[1]) + '1')
            ws.cell(row=1, column=1).value = 'Missing Value percentage'
            ws.cell(row=1, column=1).font = Font(bold=True, size=16, underline='single')
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            rows = dataframe_to_rows(df, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1, column=c_idx, value=value)
                    ws.cell(row=r_idx + 1, column=c_idx).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=r_idx + 1, column=c_idx).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            df2 = outlier_per
            for i in range(2, df2.shape[0] + 4):
                ws.row_dimensions[i + df.shape[0] + 1].height = 25
            for i in range(1, df2.shape[1] + 1):
                ws.column_dimensions[get_column_letter(i + df.shape[0] + 1)].width = 25
            for i in range(1, len(df2.columns) + 1):
                ws.cell(row=2 + df.shape[0] + 2, column=i).font = Font(bold=True)
                ws.cell(row=2 + df.shape[0] + 2, column=i).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2 + df.shape[0] + 2, column=i).fill = PatternFill("solid", fgColor="A9C4FE")
                ws.cell(row=2 + df.shape[0] + 2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            ws.merge_cells(
                'A' + str(1 + df.shape[0] + 2) + ':' + get_column_letter(df2.shape[1]) + str(1 + df.shape[0] + 2))
            ws.cell(row=1 + df.shape[0] + 2, column=1).value = 'Outlier Percentage'
            ws.cell(row=1 + df.shape[0] + 2, column=1).font = Font(bold=True, size=16, underline='single')
            ws.cell(row=1 + df.shape[0] + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            rows = dataframe_to_rows(df2, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx, value=value)
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
                    ws.cell(row=r_idx + 1 + df.shape[0] + 2, column=c_idx).border = Border(top=thin, left=thin,
                                                                                           right=thin,
                                                                                           bottom=thin)

            df3 = cramers_table
            if isinstance(df3, pd.DataFrame):
                for i in range(1, df3.shape[0] + 4):
                    ws.row_dimensions[i + df.shape[0] + df2.shape[0] + 4].height = 25
                for i in range(1, df3.shape[1] + 1):
                    ws.column_dimensions[get_column_letter(i + df2.shape[0] + 1)].width = 25
                for i in range(1, len(df3.columns) + 1):
                    ws.cell(row=2 + df.shape[0] + df2.shape[0] + 4, column=i).font = Font(bold=True)
                    ws.cell(row=2 + df.shape[0] + df2.shape[0] + 4, column=i).alignment = Alignment(horizontal='center',
                                                                                                    vertical='center')
                    ws.cell(row=2 + df.shape[0] + df2.shape[0] + 4, column=i).fill = PatternFill("solid",
                                                                                                 fgColor="A9C4FE")
                    ws.cell(row=2 + df.shape[0] + df2.shape[0] + 4, column=i).border = Border(top=thin, left=thin,
                                                                                              right=thin,
                                                                                              bottom=thin)
                ws.merge_cells(
                    'A' + str(1 + df.shape[0] + df2.shape[0] + 4) + ':' + get_column_letter(df3.shape[1]) + str(
                        1 + df.shape[0] + df2.shape[0] + 4))
                ws.cell(row=1 + df.shape[0] + df2.shape[0] + 4, column=1).value = 'Cramers Matrix'
                ws.cell(row=1 + df.shape[0] + df2.shape[0] + 4, column=1).font = Font(bold=True, size=16,
                                                                                      underline='single')
                ws.cell(row=1 + df.shape[0] + df2.shape[0] + 4, column=1).alignment = Alignment(horizontal='center',
                                                                                                vertical='center')
                rows = dataframe_to_rows(df3, index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        # if type(value) != str and type(value) != int:
                        #     value = round(value, 2)
                        ws.cell(row=r_idx + 1 + df.shape[0] + df2.shape[0] + 4, column=c_idx, value=value)
                        ws.cell(row=r_idx + 1 + df.shape[0] + df2.shape[0] + 4, column=c_idx).alignment = Alignment(
                            horizontal='center',
                            vertical='center')
                        ws.cell(row=r_idx + 1 + df.shape[0] + df2.shape[0] + 4, column=c_idx).border = Border(top=thin,
                                                                                                              left=thin,
                                                                                                              right=thin,
                                                                                                              bottom=thin)
            ws.sheet_view.showGridLines = False

        def charts(workbook, pca: pd.DataFrame, im: str, kl: np.ndarray = None, hl: np.ndarray = None,
                   gl: np.ndarray = None, bl: np.ndarray = None):
            ws = workbook.create_sheet(0)
            ws.title = 'Charts'

            plots = []
            if kl is not None:
                plots.append((kl, 'KMeans'))
            if hl is not None:
                plots.append((hl, 'Hierarchical'))
            if gl is not None:
                plots.append((gl, 'GMM'))
            if bl is not None:
                plots.append((bl, 'Birch'))

            remove = []
            endrow = 0
            for i, (labels, title) in enumerate(plots):
                fig = plt.figure()
                plt.scatter(pca['PC1'], pca['PC2'], c=labels, cmap=viridis)
                plt.title(title)
                plt.xlabel('PC1')
                plt.ylabel('PC2')
                plt.colorbar()
                plot_filename = f'{title}_plot.png'
                fig.savefig(plot_filename)
                remove.append(plot_filename)
                img = Image(plot_filename)
                img.width = 600
                img.height = 400
                row = (i // 2) * 20 + 1
                col = (i % 2) * 10 + 3
                ws.add_image(img, f'{chr(col + 64)}{row}')
                ws.sheet_view.showGridLines = False
                endrow = row
            im1 = Image(im)
            im1.width = 1250
            im1.height = 600
            ws.add_image(im1, f'{chr(67)}{endrow * 2}')
            return remove

        wb = openpyxl.Workbook()
        distribution(workbook=wb, distribution_table=self.distribution, score_table=self.score_table)
        summary(workbook=wb, org_dataset=org_dataset, dependent_variable=dependent_variable)
        describe(workbook=wb, org_dataset=org_dataset)
        analysis(workbook=wb, outlier_per=self.outlier_per, missing_table=self.missing,
                 cramers_table=self.cramers_matrix)
        remove = charts(workbook=wb, pca=pca, kl=kmeans_cluster_labels, hl=hierarchical_cluster_labels,
                        gl=gmm_cluster_labels, bl=birch_cluster_labels, im=im)
        wb.remove(wb['Sheet'])
        wb.save(filename)
        if to_delete:
            remove.append(filename)
        for i in remove:
            os.remove(i)
